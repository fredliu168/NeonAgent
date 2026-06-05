#!/usr/bin/env python3
"""Collect Fujian bidding projects from China Mobile bidding website via NeonAgent API"""
import json, urllib.request, time, re, os

BASE = "http://127.0.0.1:8787"
OUTPUT = os.path.expanduser("~/Downloads/福建招标项目.xlsx")

def call_tool(tool, args):
    req_id = f"py-{int(time.time()*1000)}"
    data = json.dumps({"requestId": req_id, "toolName": tool, "arguments": args}).encode()
    r = urllib.request.Request(f"{BASE}/tool", data=data, headers={"Content-Type":"application/json"})
    urllib.request.urlopen(r).read()
    time.sleep(3.5)
    r2 = urllib.request.urlopen(f"{BASE}/result/{req_id}")
    return json.loads(r2.read())

def get_page_content():
    resp = call_tool("read_page_content", {"selector": "body", "maxLength": 15000})
    return resp.get("data",{}).get("result","")

def click_next():
    resp = call_tool("click_element", {"selector": ".cmcc-page-next", "index": 0})
    return resp.get("data",{}).get("result","")

def parse_projects(text):
    """Parse bidding projects from page text content"""
    projects = []
    lines = text.split("\n")
    
    # Find the data rows in the table
    # Each row looks like:
    # 福建\t公告类型\t标题\t日期
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # Look for pattern: province name followed by announcement type
        if line == "福建":
            # Check next non-empty line for announcement type
            next_idx = i + 1
            while next_idx < len(lines) and lines[next_idx].strip() == "":
                next_idx += 1
            if next_idx < len(lines):
                ann_type = lines[next_idx].strip()
                if ann_type in ["采购公告", "资格预审公告", "候选人公示", "中选结果公示", "直接采购公告", "招标公告", "询比公告", "中标候选人公示", "中标结果公示", "重发公告"]:
                    # Look for title
                    title_idx = next_idx + 1
                    while title_idx < len(lines) and lines[title_idx].strip() == "":
                        title_idx += 1
                    if title_idx < len(lines):
                        title = lines[title_idx].strip()
                        # Look for date
                        date_idx = title_idx + 1
                        while date_idx < len(lines) and lines[date_idx].strip() == "":
                            date_idx += 1
                        if date_idx < len(lines):
                            date = lines[date_idx].strip()
                            date_match = re.match(r'(\d{4}-\d{2}-\d{2})', date)
                            if date_match:
                                projects.append({
                                    "region": "福建",
                                    "type": ann_type,
                                    "title": title,
                                    "date": date_match.group(1),
                                    "is_active": ann_type in ["采购公告", "招标公告", "资格预审公告", "询比公告"]
                                })
                                i = date_idx
        i += 1
    
    return projects

print("=== 开始收集福建招标项目数据 ===")

all_projects = []
unique_titles = set()

# Read page 1 (current)
print("读取第1页...")
content = get_page_content()
projects = parse_projects(content)
for p in projects:
    if p["title"] not in unique_titles:
        unique_titles.add(p["title"])
        all_projects.append(p)
print(f"  当前页: {len(projects)} 个项目, 累计: {len(all_projects)}")

# Click through pages 2-10
for page_num in range(2, 11):
    print(f"翻到第{page_num}页...")
    result = click_next()
    if "Clicked" not in result:
        print(f"  翻页失败: {result}")
        break
    time.sleep(2)
    
    content = get_page_content()
    projects = parse_projects(content)
    new_count = 0
    for p in projects:
        if p["title"] not in unique_titles:
            unique_titles.add(p["title"])
            all_projects.append(p)
            new_count += 1
    print(f"  新项目: {new_count}, 累计: {len(all_projects)}")

print(f"\n=== 共收集 {len(all_projects)} 个项目 ===")

active = [p for p in all_projects if p["is_active"]]
print(f"其中正在招标(采购公告/招标公告): {len(active)} 个")

# Save raw data to JSON
json_path = "/tmp/fujian_bidding.json"
with open(json_path, "w") as f:
    json.dump(all_projects, f, ensure_ascii=False, indent=2)
print(f"数据已保存到 {json_path}")

# Generate Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "福建招标项目"

# Headers
headers = ["序号", "项目名称", "公告类型", "发布时间", "地区", "是否正在招标"]
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
active_font = Font(bold=True, color="0070C0")

for i, p in enumerate(all_projects, 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).border = thin_border
    ws.cell(row=row, column=2, value=p["title"]).border = thin_border
    cell_type = ws.cell(row=row, column=3, value=p["type"])
    cell_type.border = thin_border
    cell_date = ws.cell(row=row, column=4, value=p["date"])
    cell_date.border = thin_border
    ws.cell(row=row, column=5, value=p["region"]).border = thin_border
    
    is_active = "是" if p["is_active"] else "否"
    cell_active = ws.cell(row=row, column=6, value=is_active)
    cell_active.border = thin_border
    
    if p["is_active"]:
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = green_fill
            ws.cell(row=row, column=col).font = active_font

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 14

# Auto filter
ws.auto_filter.ref = f"A1:F{len(all_projects)+1}"

# Freeze top row
ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Excel 已保存到: {OUTPUT}")
print(f"总项目数: {len(all_projects)}, 正在招标: {len(active)}")
