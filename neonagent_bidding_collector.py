#!/usr/bin/env python3
"""
NeonAgent + Python 集成示例脚本
通过 NeonAgent 工具 API 收集招标数据并导出为 Excel
"""
import json, urllib.request, time, re, os, sys
from datetime import datetime

# ===== 配置 =====
NEONAGENT_URL = "http://127.0.0.1:8787"
OUTPUT_FILE = os.path.expanduser("~/Downloads/福建移动招标项目.xlsx")
MAX_PAGES = 10  # 最多翻页数量

# ===== NeonAgent API 封装 =====

def call_tool(tool_name, args):
    """调用 NeonAgent 工具，等待结果后返回"""
    req_id = f"na-{int(time.time()*1000)}"
    body = json.dumps({"requestId": req_id, "toolName": tool_name, "arguments": args}).encode()
    r = urllib.request.Request(f"{NEONAGENT_URL}/tool", data=body,
                               headers={"Content-Type": "application/json"})
    urllib.request.urlopen(r).read()
    # 等待页面渲染
    time.sleep(3)
    resp = urllib.request.urlopen(f"{NEONAGENT_URL}/result/{req_id}")
    return json.loads(resp.read())

def navigate(url):
    """导航到目标网址"""
    print(f"  → 导航到: {url}")
    r = call_tool("navigate", {"url": url})
    time.sleep(2)
    return r

def read_page(max_length=15000):
    """读取页面内容"""
    r = call_tool("read_page_content", {"selector": "body", "maxLength": max_length})
    return r.get("data", {}).get("result", "")

def click(selector, index=0, text=None):
    """点击页面元素"""
    args = {"selector": selector, "index": index}
    if text:
        args["text"] = text
    r = call_tool("click_element", args)
    return r.get("data", {}).get("result", "")

def type_text(selector, text):
    """在输入框中输入文字"""
    r = call_tool("type_text", {"selector": selector, "text": text})
    return r.get("data", {}).get("result", "")

def get_page_info():
    """获取当前页面信息"""
    r = call_tool("get_page_info", {})
    info = r.get("data", {}).get("result", "{}")
    return json.loads(info)

# ===== 数据解析 =====

def parse_bidding_projects(text):
    """从页面文本中解析招标项目"""
    projects = []
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line in ["福建", "福建\n"]:
            # 查找公告类型
            j = i + 1
            while j < len(lines) and lines[j].strip() == "":
                j += 1
            if j >= len(lines): break
            ann_type = lines[j].strip()
            
            valid_types = ["采购公告", "资格预审公告", "候选人公示", "中选结果公示",
                          "直接采购公告", "招标公告", "询比公告", "重发公告",
                          "中标候选人公示", "中标结果公示"]
            if ann_type not in valid_types:
                i += 1
                continue
            
            # 查找标题
            k = j + 1
            while k < len(lines) and lines[k].strip() == "":
                k += 1
            if k >= len(lines): break
            title = lines[k].strip()
            
            # 查找日期
            l = k + 1
            while l < len(lines) and lines[l].strip() == "":
                l += 1
            if l >= len(lines): break
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', lines[l])
            
            if date_match and title:
                # 清理标题
                title = title.replace("_询比公告", "").replace("_招标公告", "")
                title = title.replace("_中选候选人公示", "").replace("_中标候选人公示", "")
                title = title.replace("_候选人公示", "").replace("_中标结果公示", "")
                title = title.replace("_中选结果公示", "").replace("_直接采购信息公告", "")
                title = title.replace("_重发公告", "")
                
                active = ann_type in ["采购公告", "招标公告", "资格预审公告", "询比公告"]
                projects.append({
                    "region": "福建",
                    "type": ann_type,
                    "title": title.strip(),
                    "date": date_match.group(1),
                    "is_active": active
                })
                i = l
        i += 1
    return projects

# ===== 主要流程 =====

def main():
    print("=" * 60)
    print("NeonAgent 招标数据采集")
    print(f"启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # Step 1: 获取当前页面信息
    print("\n[1/5] 获取当前页面信息...")
    info = get_page_info()
    print(f"  当前 URL: {info.get('url', 'unknown')}")
    print(f"  标题: {info.get('title', 'unknown')}")
    
    # Step 2: 导航到招标采购公告页面（如果不在）
    current_url = info.get("url", "")
    if "biddingProcurementBulletin" not in current_url:
        print("\n[2/5] 导航到招标公告页面...")
        navigate("https://b2b.10086.cn/#/biddingProcurementBulletin")
        time.sleep(2)
    else:
        print("\n[2/5] 已在招标公告页面 ✓")
    
    time.sleep(1)
    print(f"  当前 URL: {get_page_info().get('url', '')}")
    
    # Step 3: 搜索福建项目
    print("\n[3/5] 搜索福建项目...")
    # 先点搜索框输入"福建"
    result = type_text(".title-input input", "福建")
    print(f"  输入关键词: {result}")
    time.sleep(1)
    
    # 点击查询按钮
    result = click(".cmcc-btn-primary", index=0)
    print(f"  点击查询: {result}")
    time.sleep(3)
    
    # Step 4: 收集多页数据
    print(f"\n[4/5] 收集数据 (最多 {MAX_PAGES} 页)...")
    
    all_projects = []
    seen_titles = set()
    
    for page in range(1, MAX_PAGES + 1):
        print(f"\n  --- 第 {page} 页 ---")
        
        # 读取页面内容
        content = read_page(15000)
        
        # 打印当前页总数信息
        total_match = re.search(r'共\s*(\d+)\s*条', content)
        current_info = f"共 {total_match.group(1)} 条" if total_match else "未知"
        print(f"  页面信息: {current_info}")
        
        # 解析项目
        projects = parse_bidding_projects(content)
        new_count = 0
        for p in projects:
            if p["title"] not in seen_titles:
                seen_titles.add(p["title"])
                all_projects.append(p)
                new_count += 1
        
        active_in_page = sum(1 for p in projects if p.get("is_active"))
        print(f"  本页: {len(projects)} 项目 (其中招标中: {active_in_page})")
        print(f"  累计: {len(all_projects)} 项目 (去重)")
        
        # 翻到下一页
        if page < MAX_PAGES:
            result = click(".cmcc-page-next", index=0)
            if "Clicked" not in result:
                print(f"  翻页完成(可能已到最后一页)")
                break
            time.sleep(2)
    
    # 统计
    active_projects = [p for p in all_projects if p["is_active"]]
    print(f"\n[5/5] 生成 Excel...")
    print(f"  总项目数: {len(all_projects)}")
    print(f"  正在招标: {len(active_projects)}")
    
    # Step 5: 生成 Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "福建招标项目"
        
        # --- 表头 ---
        headers = ["序号", "项目名称", "公告类型", "发布日期", "地区", "状态"]
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # --- 数据行 ---
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        active_font = Font(bold=True, color="006100")
        normal_font = Font(color="333333")
        
        for i, p in enumerate(all_projects, 1):
            row = i + 1
            vals = [i, p["title"], p["type"], p["date"], p["region"],
                    "● 正在招标" if p["is_active"] else "已结束"]
            
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 2))
                
                if p["is_active"]:
                    cell.fill = green_fill
                    cell.font = active_font
                else:
                    cell.font = normal_font
        
        # --- 格式 ---
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        
        # 自动筛选 + 冻结窗格
        ws.auto_filter.ref = f"A1:F{len(all_projects)+1}"
        ws.freeze_panes = "A2"
        
        # 设置行高
        ws.row_dimensions[1].height = 25
        for i in range(2, len(all_projects) + 2):
            ws.row_dimensions[i].height = 30
        
        wb.save(OUTPUT_FILE)
        print(f"\n  ✅ Excel 已保存: {OUTPUT_FILE}")
        
    except ImportError:
        print("  ⚠️ openpyxl 未安装，输出为 JSON 文件")
        json_file = OUTPUT_FILE.replace(".xlsx", ".json")
        with open(json_file, "w") as f:
            json.dump(all_projects, f, ensure_ascii=False, indent=2)
        print(f"  📄 JSON 已保存: {json_file}")
    
    # 打印活跃项目摘要
    print(f"\n📋 正在招标项目摘要:")
    for i, p in enumerate(active_projects, 1):
        print(f"  {i:2d}. [{p['date']}] {p['title'][:60]}")
    
    print(f"\n{'=' * 60}")
    print(f"采集完成! 共 {len(all_projects)} 个项目")
    print(f"正在招标: {len(active_projects)} 个")
    print(f"{'=' * 60}")
    
    return all_projects

if __name__ == "__main__":
    main()
